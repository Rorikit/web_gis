from datetime import date

from django.contrib.auth import get_user_model
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase

from apps.accounts.models import District, UserRole
from apps.incidents.models import AreaState, ContractorType, Damage, DamageType, NetworkType, OrderKind

User = get_user_model()


class IncidentsApiTests(APITestCase):
    def setUp(self):
        self.central = District.objects.create(id='central', name='Центральный район')
        self.north = District.objects.create(id='north', name='Северный район')

        self.admin = User.objects.create_user(username='admin', password='admin', is_superuser=True, is_staff=True)
        self.admin.profile.role = UserRole.ADMIN
        self.admin.profile.full_name = 'Администратор'
        self.admin.profile.save()

        self.damage_user = User.objects.create_user(username='ivanov', password='ivanov')
        self.damage_user.profile.role = UserRole.DISTRICT_DAMAGE
        self.damage_user.profile.full_name = 'Иванов'
        self.damage_user.profile.district = self.central
        self.damage_user.profile.save()

        self.order_user = User.objects.create_user(username='sidorov', password='sidorov')
        self.order_user.profile.role = UserRole.DISTRICT_ORDER
        self.order_user.profile.full_name = 'Сидоров'
        self.order_user.profile.district = self.north
        self.order_user.profile.save()

        self.oopppr_user = User.objects.create_user(username='petrova', password='petrova')
        self.oopppr_user.profile.role = UserRole.OOPPPR
        self.oopppr_user.profile.full_name = 'Петрова'
        self.oopppr_user.profile.save()

        self.damage_central = Damage.objects.create(
            district=self.central,
            address='ул. Ленина, 12',
            network_type=NetworkType.OT,
            detected_at=date(2026, 5, 11),
            order_number='ОР-2026-0012',
            order_opened_at=date(2026, 5, 11),
            order_valid_until=date(2026, 6, 10),
            heat_source='ТЭЦ-3',
            damage_type=DamageType.CURRENT,
            disconnected_addresses='ул. Ленина, 10',
            damage_description='Повреждение трубопровода',
            order_kind=OrderKind.CURRENT,
            green_zone_area=12,
            asphalt_area=8,
            improvement_main=True,
            improvement_inner_road=False,
            improvement_sidewalk=True,
            improvement_blind_area=False,
            curb_count=3,
            area_state=AreaState.IN_PROGRESS,
            contractor_type=ContractorType.CONTRACTOR,
            contract_number='Д-44/26',
            planned_finish_date=date(2026, 5, 28),
            note='Контроль восстановления',
            archived=False,
        )

        self.damage_north = Damage.objects.create(
            district=self.north,
            address='пр. Мира, 78',
            network_type=NetworkType.GVS,
            detected_at=date(2026, 5, 15),
            fixed_at=date(2026, 5, 17),
            order_number='ОР-2026-0031',
            order_opened_at=date(2026, 5, 15),
            order_valid_until=date(2026, 6, 14),
            heat_source='Котельная Северная',
            damage_type=DamageType.HYDRAULIC,
            disconnected_addresses='',
            damage_description='Повреждение после испытаний',
            order_kind=OrderKind.GUARANTEE,
            green_zone_area=0,
            asphalt_area=18,
            improvement_main=True,
            improvement_inner_road=True,
            improvement_sidewalk=False,
            improvement_blind_area=False,
            curb_count=0,
            area_state=AreaState.READY_TO_CLOSE,
            contractor_type=ContractorType.URTS,
            contract_number='',
            planned_finish_date=date(2026, 5, 25),
            archived=False,
        )

        self.damage_without_order = Damage.objects.create(
            district=self.central,
            address='ул. Свободы, 5',
            network_type=NetworkType.OT,
            detected_at=date(2026, 5, 20),
            heat_source='ТЭЦ-3',
            damage_type=DamageType.CURRENT,
            damage_description='Повреждение без ордера',
            area_state=AreaState.IN_PROGRESS,
            contractor_type=ContractorType.CONTRACTOR,
            archived=False,
        )

    def test_damage_without_order_excluded_from_orders_list(self):
        self.client.login(username='admin', password='admin')

        response = self.client.get('/orders', {'archived': 'false'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        ids = [item['id'] for item in response.json()]
        self.assertNotIn(self.damage_without_order.id, ids)
        self.assertIn(self.damage_central.id, ids)

    def test_open_order_creates_order_fields_and_surfaces_in_orders_list(self):
        self.client.login(username='ivanov', password='ivanov')

        response = self.client.post(f'/damages/{self.damage_without_order.id}/open-order')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        body = response.json()
        self.assertTrue(body['orderNumber'])
        self.assertIsNotNone(body['orderOpenedAt'])
        self.assertEqual(body['orderKind'], OrderKind.CURRENT)

        orders_response = self.client.get('/orders', {'archived': 'false'})
        ids = [item['id'] for item in orders_response.json()]
        self.assertIn(self.damage_without_order.id, ids)

        # opening twice is rejected
        again = self.client.post(f'/damages/{self.damage_without_order.id}/open-order')
        self.assertEqual(again.status_code, status.HTTP_400_BAD_REQUEST)

    def test_damage_list_filtered_by_district(self):
        self.client.login(username='ivanov', password='ivanov')

        response = self.client.get('/damages', {'archived': 'false'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        body = response.json()
        ids = {item['id'] for item in body}
        self.assertEqual(ids, {self.damage_central.id, self.damage_without_order.id})

    def test_damage_update_allowed_for_order_role_same_district(self):
        # ТЗ раздел Б: техник по повреждениям и техник по ордерам одного района имеют
        # одинаковый набор прав на поля повреждения/ордера.
        self.client.login(username='sidorov', password='sidorov')

        response = self.client.put(
            f'/damages/{self.damage_north.id}',
            {'note': 'new note'},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.damage_north.refresh_from_db()
        self.assertEqual(self.damage_north.note, 'new note')

    def test_damage_update_forbidden_for_foreign_district_order_role(self):
        self.client.login(username='sidorov', password='sidorov')

        response = self.client.put(
            f'/damages/{self.damage_central.id}',
            {'note': 'new note'},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_damage_create_forbidden_for_foreign_district(self):
        self.client.login(username='ivanov', password='ivanov')

        response = self.client.post(
            '/damages',
            {'districtId': self.north.id, 'address': 'пр. Победы, 1'},
            format='json',
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_damage_update_forbidden_to_move_into_foreign_district(self):
        self.client.login(username='ivanov', password='ivanov')

        response = self.client.put(
            f'/damages/{self.damage_central.id}',
            {'districtId': self.north.id},
            format='json',
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_order_update_oopppr_only_oopppr_fields(self):
        self.client.login(username='petrova', password='petrova')

        response = self.client.put(
            f'/orders/{self.damage_central.id}',
            {
                'contractorType': 'УРТС',
                'contractNumber': 'K-100',
                'plannedFinishDate': '2026-06-01',
                'note': 'ok',
                'address': 'Адрес не должен обновиться',
            },
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.damage_central.refresh_from_db()
        self.assertEqual(self.damage_central.contractor_type, 'УРТС')
        self.assertEqual(self.damage_central.address, 'ул. Ленина, 12')

    def test_gis_save_point(self):
        self.client.login(username='ivanov', password='ivanov')

        response = self.client.post(
            f'/gis/damages/{self.damage_central.id}/point',
            {'latitude': 55.7558, 'longitude': 37.6173},
            format='json',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        body = response.json()
        self.assertIsNotNone(body['gisPoint'])
        self.assertEqual(body['gisPoint']['latitude'], 55.7558)

    def test_users_list_admin_only(self):
        self.client.login(username='ivanov', password='ivanov')
        denied = self.client.get('/users')
        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)

        self.client.logout()
        self.client.login(username='admin', password='admin')
        allowed = self.client.get('/users')
        self.assertEqual(allowed.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(len(allowed.json()), 4)

    def test_users_create_admin_only(self):
        self.client.login(username='ivanov', password='ivanov')
        denied = self.client.post(
            '/users',
            {'ldapLogin': 'newuser', 'password': 'Str0ng!Pass123', 'role': UserRole.DISTRICT_DAMAGE},
            format='json',
        )
        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)

    def test_users_create_success(self):
        self.client.login(username='admin', password='admin')
        response = self.client.post(
            '/users',
            {
                'ldapLogin': 'newuser',
                'password': 'Str0ng!Pass123',
                'fullName': 'Новый Пользователь',
                'role': UserRole.DISTRICT_ORDER,
                'districtId': self.north.id,
                'isActive': True,
            },
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        body = response.json()
        self.assertEqual(body['ldapLogin'], 'newuser')
        self.assertEqual(body['role'], UserRole.DISTRICT_ORDER)
        self.assertEqual(body['districtId'], self.north.id)

        created_user = User.objects.get(username='newuser')
        self.assertTrue(created_user.check_password('Str0ng!Pass123'))

    def test_users_create_duplicate_login(self):
        self.client.login(username='admin', password='admin')
        response = self.client.post(
            '/users',
            {'ldapLogin': 'ivanov', 'password': 'Str0ng!Pass123', 'role': UserRole.DISTRICT_DAMAGE},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_409_CONFLICT)

    def test_users_create_weak_password_rejected(self):
        self.client.login(username='admin', password='admin')
        response = self.client.post(
            '/users',
            {'ldapLogin': 'weakuser', 'password': '123', 'role': UserRole.DISTRICT_DAMAGE},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(User.objects.filter(username='weakuser').exists())

    def test_export_current_table_damages(self):
        self.client.login(username='admin', password='admin')
        response = self.client.post('/exports/current-table', {'entityType': 'damages', 'archived': False}, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue(len(response.content) > 0)

    def test_export_current_table_scoped_to_district(self):
        self.client.login(username='ivanov', password='ivanov')
        response = self.client.post('/exports/current-table', {'entityType': 'damages', 'archived': False}, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_audit_history(self):
        self.client.login(username='ivanov', password='ivanov')
        self.client.put(f'/damages/{self.damage_central.id}', {'note': 'changed'}, format='json')

        response = self.client.get(f'/audit/damage/{self.damage_central.id}')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(len(response.json()), 1)

    def test_order_update_district_role_can_edit_order_dates(self):
        self.client.login(username='ivanov', password='ivanov')

        response = self.client.put(
            f'/orders/{self.damage_central.id}',
            {
                'openedAt': '2026-05-12',
                'validUntil': '2026-07-01',
                'closedAt': '2026-06-20',
                'contractorRequestDate': '2026-05-15',
            },
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        body = response.json()
        self.assertEqual(body['openedAt'], '2026-05-12')
        self.assertEqual(body['validUntil'], '2026-07-01')
        self.assertEqual(body['closedAt'], '2026-06-20')
        self.assertEqual(body['contractorRequestDate'], '2026-05-15')

    def test_order_update_oopppr_cannot_edit_order_dates(self):
        self.client.login(username='petrova', password='petrova')

        response = self.client.put(
            f'/orders/{self.damage_central.id}',
            {'openedAt': '2020-01-01', 'closedAt': '2020-02-02', 'note': 'ООППР'},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.damage_central.refresh_from_db()
        self.assertEqual(self.damage_central.order_opened_at, date(2026, 5, 11))
        self.assertIsNone(self.damage_central.order_closed_at)
        self.assertEqual(self.damage_central.note, 'ООППР')

    def test_order_serializer_exposes_appendix5_fields(self):
        self.client.login(username='admin', password='admin')
        response = self.client.get(f'/orders/{self.damage_central.id}')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        body = response.json()
        for field in (
            'orderNumber', 'address', 'orderKind', 'openedAt', 'validUntil', 'closedAt',
            'greenZoneArea', 'asphaltArea', 'improvementMain', 'improvementInnerRoad',
            'improvementSidewalk', 'improvementBlindArea', 'curbCount', 'areaState',
            'contractorType', 'contractNumber', 'contractorRequestDate', 'plannedFinishDate',
            'note', 'photos', 'gisPoint',
        ):
            self.assertIn(field, body)

    def test_reference_report_counts_orders_by_district(self):
        from openpyxl import load_workbook
        import io

        Damage.objects.filter(pk=self.damage_north.pk).update(
            order_kind=OrderKind.GUARANTEE,
            order_opened_at=date(2026, 1, 5),
            order_closed_at=date(2026, 3, 3),
        )

        self.client.login(username='petrova', password='petrova')
        response = self.client.post('/reports/reference', {'reportDate': '2026-06-01'}, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        sheet = load_workbook(io.BytesIO(response.content)).active
        rows = {row[0]: row[1:] for row in sheet.iter_rows(min_row=6, values_only=True)}

        # Центральный: ордер открыт 11.05.2026 и не закрыт -> «Открыто/Текущие» + «в работе»
        self.assertEqual(rows['Центральный район'][:6], (0, 0, 1, 0, 1, 0))
        # Северный: гарантийный ордер закрыт 03.03.2026 -> «Закрыто/Гарантийные»
        self.assertEqual(rows['Северный район'][:6], (0, 1, 0, 0, 0, 0))

    def test_damage_card_report_uses_appendix2_fields(self):
        self.client.login(username='ivanov', password='ivanov')
        response = self.client.post(
            '/reports/damage-card',
            {'damageId': self.damage_central.id, 'fields': {'startChamber': 'ТК-15', 'unknownField': 'x'}},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(len(response.content) > 0)

    def test_order_close_date_moves_order_to_archive(self):
        self.client.login(username='ivanov', password='ivanov')

        self.client.put(f'/orders/{self.damage_central.id}', {'closedAt': '2026-06-20'}, format='json')
        self.damage_central.refresh_from_db()
        self.assertTrue(self.damage_central.archived)

        self.client.put(f'/orders/{self.damage_central.id}', {'closedAt': None}, format='json')
        self.damage_central.refresh_from_db()
        self.assertFalse(self.damage_central.archived)
